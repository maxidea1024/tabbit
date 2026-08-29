# -*- coding: utf-8 -*-
"""test/fixtures/xlsx/matrix/matrix.xlsx 를 다시 쓴다.

`:matrix` 선언이 읽는 격자를 담습니다 — 축이 정수인 것, 축이 enum인 것, 그리고
메모 컬럼과 제외 행. 고치면 이 파일을 고치고 다시 실행한 뒤 골든을 재기록합니다.

    python test/fixtures/matrix-workbook.py
"""
import os

import openpyxl

OUT = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "xlsx", "matrix", "matrix.xlsx")

# 시트마다 (B2부터 놓을 행들). 첫 행과 첫 열을 비우는 것은 다른 픽스처와 같습니다 —
# 레이아웃이 시트의 어디에서든 선언을 찾는다는 것을 픽스처가 계속 확인합니다.
SHEETS = {
    "Enums": [
        [":enum Element", "What an attack is made of."],
        [":field", "label", "value", "desc"],
        ["", "None", "0", "no element"],
        ["", "Fire", "1", "burns"],
        ["", "Water", "2", "soaks"],
        ["", "Wind", "3", "cuts"],
    ],
    "Goods": [
        [":table Goods", "What a town trades in."],
        [":field", "id", "name"],
        [":type", "int", "string"],
        ["", "101", "salt"],
        ["", "102", "silk"],
        ["", "103", "tea"],
    ],
    "Grids": [
        # 축이 정수이고 다른 테이블을 가리킵니다. `#` 컬럼은 메모, `-` 는 값 없음,
        # 마커 열의 `#` 는 그 행 하나를 뺍니다.
        [":matrix TownPrice", "Price modifier by town and goods."],
        [":field", "town", "price"],
        [":type", "int", "int? (min=-200, max=200)"],
        [":desc", "which town", "the modifier"],
        [":col", "goods foreign Goods", "101", "102", "#", "103"],
        ["", "2001", "0", "-25", "check", "-125"],
        ["", "2002", "10", "-", "check", "-40"],
        ["#", "2003", "1", "2", "left out", "3"],
        [],
        # 두 축이 같은 enum입니다. 열의 키가 라벨이라 판독 규칙으로는 적을 수 없는 표이고,
        # 선언이 있어서 되는 것이 이것입니다.
        [":matrix ElementChart", "How much an element does to another."],
        [":field", "attacker", "rate"],
        [":type", "Element", "float"],
        [":col", "defender Element", "Fire", "Water", "Wind"],
        ["", "Fire", "1.0", "0.5", "2.0"],
        ["", "Water", "2.0", "1.0", "1.0"],
        ["", "Wind", "0.5", "1.0", "1.0"],
    ],
}


def build():
    book = openpyxl.Workbook()
    book.remove(book.active)

    for title, rows in SHEETS.items():
        sheet = book.create_sheet(title)

        for r, row in enumerate(rows):
            for c, value in enumerate(row):
                if value == "":
                    continue
                sheet.cell(row=r + 2, column=c + 2, value=value)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    book.save(OUT)
    print(OUT)


if __name__ == "__main__":
    build()
